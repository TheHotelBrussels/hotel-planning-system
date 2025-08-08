import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import streamlit as st
from pulp import LpProblem, LpMinimize, LpVariable, lpSum, LpStatus
import json
from dataclasses import dataclass
from typing import List, Dict, Optional
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.styles

# ================================
# MODELE DE DONNEES
# ================================

@dataclass
class Employee:
    prenom: str
    nom: str
    type_contrat: str
    jours_semaine: int
    role: str
    contraintes_speciales: Dict
    competences: List[str]
    disponible: bool = True  # Nouveau : disponibilité
    motif_indisponibilite: str = ""  # Nouveau : raison si indisponible
    jours_absence: int = 0  # Nouveau : nombre de jours d'absence dans la semaine (0-7)
    jours_off_consecutifs: Optional[bool] = None

    def __post_init__(self):
        if self.type_contrat == 'temps_plein':
            self.jours_semaine = 5
        elif self.type_contrat == 'mi_temps_4j':
            self.jours_semaine = 4
        elif self.type_contrat == 'mi_temps_3j':
            self.jours_semaine = 3
        elif self.type_contrat == 'nuit':
            self.jours_semaine = 5

    @property
    def jours_travail_max_semaine(self):
        """Calcule le nombre maximum de jours de travail possible cette semaine"""
        if not self.disponible:
            return 0
        elif self.jours_absence >= 7:
            return 0
        else:
            # Jours contractuels moins les jours d'absence, minimum 0
            return max(0, self.jours_semaine - self.jours_absence)

# ================================
# SYSTEME DE PLANNING
# ================================

class HotelPlanningSystem:
    def __init__(self):
        self.employees = []
        self.hotel_capacity = 422
        self.clients_per_receptionist = 50
        self.max_receptionists_per_shift = 4
        self.nb_supervisors = 5  # 5 superviseurs qui font aussi réceptionnistes
        self.nb_receptionnistes_jour = 6  # 6 réceptionnistes jour classiques
        self.nb_night_receptionists_total = 3
        self.nb_night_receptionists_required = 2
        self.nb_concierges = 1
        self.jours_semaine = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']

    def ajouter_employe(self, employee: Employee):
        self.employees.append(employee)

    def supprimer_employe(self, prenom, nom):
        self.employees = [e for e in self.employees if not (e.prenom == prenom and e.nom == nom)]

    def calculer_besoins_personnel(self, checkins: Dict[str, int], checkouts: Dict[str, int]) -> Dict:
        besoins = {}
        for jour in self.jours_semaine:
            nb_checkins = checkins.get(jour, 0)
            nb_checkouts = checkouts.get(jour, 0)

            # Calcul des besoins totaux en personnel (superviseurs + réceptionnistes)
            besoin_total_matin = max(1, int(np.ceil(nb_checkouts / self.clients_per_receptionist)))
            besoin_total_apres_midi = max(1, int(np.ceil(nb_checkins / self.clients_per_receptionist)))

            # Maximum 4 personnes par shift (incluant concierge si présent)
            max_personnel_disponible = self.max_receptionists_per_shift
            if jour not in ['Samedi', 'Dimanche']:
                max_personnel_disponible -= 1  # -1 pour le concierge en semaine

            besoin_total_matin = min(besoin_total_matin, max_personnel_disponible)
            besoin_total_apres_midi = min(besoin_total_apres_midi, max_personnel_disponible)

            # Ajustement selon l'occupation - si faible, on peut réduire les besoins
            total_activite = nb_checkins + nb_checkouts
            if total_activite < 100:  # Occupation faible
                besoin_total_matin = max(1, besoin_total_matin - 1)
                besoin_total_apres_midi = max(1, besoin_total_apres_midi - 1)

            besoins[jour] = {
                'matin': {
                    'total_personnel': besoin_total_matin,  # Total superviseurs + réceptionnistes
                    'min_superviseurs': 1,  # Au moins 1 superviseur obligatoire
                    'concierge': 1 if jour not in ['Samedi', 'Dimanche'] else 0
                },
                'apres_midi': {
                    'total_personnel': besoin_total_apres_midi,
                    'min_superviseurs': 1,  # Au moins 1 superviseur obligatoire
                    'concierge': 0  # Concierge uniquement le matin
                },
                'nuit': {
                    'receptionists': min(self.nb_night_receptionists_required, 
                                       len([e for e in self.employees if e.role == 'receptionniste' 
                                           and e.type_contrat == 'nuit' and e.disponible])),
                    'superviseurs': 0,
                    'concierge': 0
                }
            }
        return besoins

    def get_employees_disponibles(self) -> List[Employee]:
        """Retourne uniquement les employés disponibles"""
        return [emp for emp in self.employees if emp.disponible]

    def verifier_faisabilite_planning(self, besoins: Dict) -> Dict:
        """Vérifie si le planning est réalisable avec l'équipe disponible"""
        employes_disponibles = self.get_employees_disponibles()
        superviseurs_dispo = [e for e in employes_disponibles if e.role == 'superviseur']
        receptionnistes_jour_dispo = [e for e in employes_disponibles if e.role == 'receptionniste' and e.type_contrat != 'nuit']
        receptionnistes_nuit_dispo = [e for e in employes_disponibles if e.role == 'receptionniste' and e.type_contrat == 'nuit']
        concierges_dispo = [e for e in employes_disponibles if e.role == 'concierge']

        problemes = []
        recommandations = []

        # Vérifications critiques
        if len(superviseurs_dispo) < 1:
            problemes.append("❌ CRITIQUE: Aucun superviseur disponible")
        elif len(superviseurs_dispo) < 2:
            recommandations.append("⚠️ Un seul superviseur disponible - couverture limitée")

        if len(receptionnistes_nuit_dispo) < 2:
            if len(receptionnistes_nuit_dispo) == 0:
                problemes.append("❌ CRITIQUE: Aucun réceptionniste de nuit disponible")
            else:
                problemes.append("❌ CRITIQUE: Un seul réceptionniste de nuit disponible (2 requis)")

        if len(concierges_dispo) == 0:
            recommandations.append("⚠️ Concierge indisponible - service limité en semaine")

        # Calcul de la charge de travail
        total_personnel_jour = len(superviseurs_dispo) + len(receptionnistes_jour_dispo)
        if total_personnel_jour < 3:
            problemes.append("❌ Personnel jour insuffisant (minimum 3 pour couvrir les shifts)")

        return {
            'faisable': len(problemes) == 0,
            'problemes': problemes,
            'recommandations': recommandations,
            'stats': {
                'total_disponibles': len(employes_disponibles),
                'superviseurs': len(superviseurs_dispo),
                'receptionnistes_jour': len(receptionnistes_jour_dispo),
                'receptionnistes_nuit': len(receptionnistes_nuit_dispo),
                'concierges': len(concierges_dispo)
            }
        }

    def generer_planning_optimise(self, checkins: Dict[str, int], checkouts: Dict[str, int], semaine_debut: datetime) -> Dict:
        besoins = self.calculer_besoins_personnel(checkins, checkouts)
        prob = LpProblem("Planning_Front_Office", LpMinimize)

        # Variables décision
        x = {}
        for emp in self.employees:
            x[emp.prenom] = {}
            for jour in self.jours_semaine:
                x[emp.prenom][jour] = {}
                for shift in ['matin', 'apres_midi', 'nuit']:
                    x[emp.prenom][jour][shift] = LpVariable(f"x_{emp.prenom}_{jour}_{shift}", cat='Binary')

        # Fonction objectif : minimiser le nombre total d'affectations
        prob += lpSum([
            x[emp.prenom][jour][shift]
            for emp in self.employees
            for jour in self.jours_semaine
            for shift in ['matin', 'apres_midi', 'nuit']
        ])

        # Ajouter toutes les contraintes
        self._ajouter_contraintes(prob, x, besoins)
        
        # Résoudre
        prob.solve()

        if prob.status != 1:
            st.warning(f"⚠️ Statut du solveur : {LpStatus[prob.status]}. Solution approchée proposée.")

        return self._extraire_planning(x)

    def _ajouter_contraintes(self, prob, x, besoins):
        # Listes des employés par type - SEULEMENT LES DISPONIBLES
        employes_disponibles = self.get_employees_disponibles()
        superviseurs = [e for e in employes_disponibles if e.role == 'superviseur']
        receptionnistes_jour = [e for e in employes_disponibles if e.role == 'receptionniste' and e.type_contrat != 'nuit']
        receptionnistes_nuit = [e for e in employes_disponibles if e.role == 'receptionniste' and e.type_contrat == 'nuit']
        concierges = [e for e in employes_disponibles if e.role == 'concierge']

        # Contrainte : les employés indisponibles ne peuvent pas être assignés
        for emp in self.employees:
            if not emp.disponible:
                for jour in self.jours_semaine:
                    for shift in ['matin', 'apres_midi', 'nuit']:
                        prob += x[emp.prenom][jour][shift] == 0

        # Contraintes de couverture par shift
        for jour in self.jours_semaine:
            for shift in ['matin', 'apres_midi', 'nuit']:
                if shift == 'nuit':
                    # Réceptionnistes de nuit selon disponibilité
                    nb_requis = min(besoins[jour][shift]['receptionists'], len(receptionnistes_nuit))
                    if nb_requis > 0:
                        prob += lpSum([x[e.prenom][jour][shift] for e in receptionnistes_nuit]) == nb_requis
                    
                    # Aucun autre type d'employé la nuit
                    autres = [e for e in employes_disponibles if e not in receptionnistes_nuit]
                    for emp in autres:
                        prob += x[emp.prenom][jour][shift] == 0
                else:
                    # Shifts jour : au moins 1 superviseur si disponible
                    if len(superviseurs) > 0:
                        prob += lpSum([x[s.prenom][jour][shift] for s in superviseurs]) >= 1

                    # Nombre total de personnel selon les besoins et disponibilité
                    nb_besoin = besoins[jour][shift]['total_personnel']
                    personnel_jour_disponible = superviseurs + receptionnistes_jour
                    nb_possible = min(nb_besoin, len(personnel_jour_disponible))
                    
                    if nb_possible > 0:
                        prob += (
                            lpSum([x[e.prenom][jour][shift] for e in receptionnistes_jour]) +
                            lpSum([x[s.prenom][jour][shift] for s in superviseurs])
                        ) >= nb_possible

                    # Concierge : selon disponibilité
                    if jour not in ['Samedi', 'Dimanche'] and shift == 'matin' and len(concierges) > 0:
                        prob += lpSum([x[c.prenom][jour][shift] for c in concierges]) == 1
                    else:
                        for c in concierges:
                            prob += x[c.prenom][jour][shift] == 0

                    # Maximum 4 personnes par shift
                    tous_employes_jour = personnel_jour_disponible + concierges
                    if len(tous_employes_jour) > 0:
                        prob += lpSum([x[e.prenom][jour][shift] for e in tous_employes_jour]) <= self.max_receptionists_per_shift

        # Contraintes par employé - SEULEMENT LES DISPONIBLES
        for emp in employes_disponibles:
            # Un seul shift par jour maximum
            for jour in self.jours_semaine:
                prob += lpSum([x[emp.prenom][jour][shift] for shift in ['matin', 'apres_midi', 'nuit']]) <= 1

            # Respect du nombre de jours de travail disponibles (contractuels - absences)
            jours_max_cette_semaine = emp.jours_travail_max_semaine
            prob += lpSum([
                lpSum([x[emp.prenom][jour][shift] for shift in ['matin', 'apres_midi', 'nuit']])
                for jour in self.jours_semaine
            ]) <= jours_max_cette_semaine

            # Contrainte : maximum 5 jours consécutifs de travail
            for i in range(len(self.jours_semaine) - 5):
                jours_seq = self.jours_semaine[i:i+6]
                prob += lpSum([
                    lpSum([x[emp.prenom][j][shift] for shift in ['matin', 'apres_midi', 'nuit']])  
                    for j in jours_seq
                ]) <= 5

            # Contraintes spécifiques par rôle
            if emp.role == 'concierge':
                # Concierge uniquement le matin en semaine
                for jour in self.jours_semaine:
                    prob += x[emp.prenom][jour]['nuit'] == 0
                    prob += x[emp.prenom][jour]['apres_midi'] == 0
                    if jour in ['Samedi', 'Dimanche']:
                        prob += x[emp.prenom][jour]['matin'] == 0

    def _extraire_planning(self, x) -> Dict:
        planning = {}
        for jour in self.jours_semaine:
            planning[jour] = {'matin': [], 'apres_midi': [], 'nuit': []}
            for emp in self.employees:
                for shift in ['matin', 'apres_midi', 'nuit']:
                    if x[emp.prenom][jour][shift].varValue == 1:
                        planning[jour][shift].append({
                            'prenom': emp.prenom,
                            'nom': emp.nom,
                            'role': emp.role,
                            'type_contrat': emp.type_contrat
                        })
        return planning
    
    def analyser_planning(self, planning: Dict) -> Dict:
        """Analyse complète du planning généré"""
        analyse = {
            'heures_par_employe': {},
            'couverture_par_shift': {},
            'violations_contraintes': [],
            'statistiques_globales': {}
        }

        # Heures et jours travaillés par employé
        for emp in self.employees:
            heures = 0
            jours_travailles = 0
            shifts_travailles = []
            
            for jour in self.jours_semaine:
                jour_travaille = False
                for shift in ['matin', 'apres_midi', 'nuit']:
                    if any(e['prenom'] == emp.prenom for e in planning[jour][shift]):
                        if shift == 'nuit':
                            heures += 8  # 8h de nuit
                        else:
                            heures += 8  # 8h de jour
                        jour_travaille = True
                        shifts_travailles.append(f"{jour}_{shift}")
                        
                if jour_travaille:
                    jours_travailles += 1
                    
            analyse['heures_par_employe'][f"{emp.prenom} {emp.nom}"] = {
                'heures': heures,
                'jours_travailles': jours_travailles,
                'jours_contractuels': emp.jours_semaine,
                'respect_contrat': jours_travailles <= emp.jours_semaine,
                'role': emp.role,
                'type_contrat': emp.type_contrat,
                'shifts': shifts_travailles
            }

        # Couverture par shift
        for jour in self.jours_semaine:
            for shift in ['matin', 'apres_midi', 'nuit']:
                equipe = planning[jour][shift]
                nb_receptionists = len([e for e in equipe if e['role'] == 'receptionniste'])
                nb_superviseurs = len([e for e in equipe if e['role'] == 'superviseur'])
                nb_concierges = len([e for e in equipe if e['role'] == 'concierge'])
                
                analyse['couverture_par_shift'][f"{jour}_{shift}"] = {
                    'total': len(equipe),
                    'receptionnistes': nb_receptionists,
                    'superviseurs': nb_superviseurs,
                    'concierge': nb_concierges,
                    'equipe': [f"{e['prenom']} {e['nom']} ({e['role']})" for e in equipe]
                }

        # Vérification des contraintes
        analyse['violations_contraintes'] = self._verifier_violations(planning)

        # Statistiques globales
        total_shifts = sum(len(planning[jour][shift]) for jour in self.jours_semaine for shift in ['matin', 'apres_midi', 'nuit'])
        total_heures = total_shifts * 8
        
        analyse['statistiques_globales'] = {
            'total_shifts_semaine': total_shifts,
            'total_heures_semaine': total_heures,
            'nombre_employes_actifs': len([emp for emp, data in analyse['heures_par_employe'].items() if data['jours_travailles'] > 0])
        }

        return analyse

    def _verifier_violations(self, planning: Dict) -> List[str]:
        violations = []
        
        for jour in self.jours_semaine:
            # Vérifications pour les shifts jour
            for shift in ['matin', 'apres_midi']:
                equipe = planning[jour][shift]
                nb_superviseurs = len([e for e in equipe if e['role'] == 'superviseur'])
                nb_concierges = len([e for e in equipe if e['role'] == 'concierge'])
                nb_receptionnistes = len([e for e in equipe if e['role'] == 'receptionniste'])
                
                # Au moins 1 superviseur obligatoire
                if nb_superviseurs < 1:
                    violations.append(f"VIOLATION: {jour} {shift} - Doit avoir au moins 1 superviseur (trouvé: {nb_superviseurs})")

                # Concierge en semaine le matin uniquement
                if jour not in ['Samedi', 'Dimanche']:
                    if shift == 'matin' and nb_concierges != 1:
                        violations.append(f"VIOLATION: {jour} {shift} - Concierge obligatoire en semaine le matin (trouvé: {nb_concierges})")
                    elif shift == 'apres_midi' and nb_concierges > 0:
                        violations.append(f"VIOLATION: {jour} {shift} - Concierge interdit l'après-midi")
                else:
                    if nb_concierges > 0:
                        violations.append(f"VIOLATION: {jour} {shift} - Concierge interdit le weekend")

                # Maximum 4 personnes par shift
                total = len(equipe)
                if total > self.max_receptionists_per_shift:
                    violations.append(f"VIOLATION: {jour} {shift} - Maximum {self.max_receptionists_per_shift} personnes (trouvé: {total})")

                # Minimum 1 personne par shift
                if total < 1:
                    violations.append(f"VIOLATION: {jour} {shift} - Au moins 1 personne requise (trouvé: {total})")

            # Vérifications pour la nuit
            equipe_nuit = planning[jour]['nuit']
            nb_receptionists_nuit = len([e for e in equipe_nuit if e['role'] == 'receptionniste'])
            nb_superviseurs_nuit = len([e for e in equipe_nuit if e['role'] == 'superviseur'])
            nb_concierges_nuit = len([e for e in equipe_nuit if e['role'] == 'concierge'])
            
            if nb_receptionists_nuit != self.nb_night_receptionists_required:
                violations.append(f"VIOLATION: {jour} nuit - Doit avoir exactement {self.nb_night_receptionists_required} réceptionnistes (trouvé: {nb_receptionists_nuit})")

            if nb_superviseurs_nuit > 0:
                violations.append(f"VIOLATION: {jour} nuit - Aucun superviseur autorisé la nuit")

            if nb_concierges_nuit > 0:
                violations.append(f"VIOLATION: {jour} nuit - Aucun concierge autorisé la nuit")

        # Vérifications des contraintes employés
        for emp in self.employees:
            jours_travailles = 0
            
            for jour in self.jours_semaine:
                travaille_ce_jour = False
                for shift in ['matin', 'apres_midi', 'nuit']:
                    if any(e['prenom'] == emp.prenom for e in planning[jour][shift]):
                        travaille_ce_jour = True
                        break
                if travaille_ce_jour:
                    jours_travailles += 1

            # Vérification du respect du contrat
            if jours_travailles > emp.jours_semaine:
                violations.append(f"VIOLATION: {emp.prenom} {emp.nom} - Travaille {jours_travailles} jours au lieu de {emp.jours_semaine} max")

        return violations

    def exporter_planning_excel(self, planning: Dict, analyse: Dict, semaine_debut: datetime) -> bytes:
        """Exporte le planning au format Excel avec mise en forme"""
        wb = Workbook()
        
        # Couleurs pour les rôles
        couleurs_roles = {
            'superviseur': PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),
            'receptionniste': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
            'concierge': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        }
        
        # Feuille 1: Planning Principal (Format Tableau avec sous-colonnes)
        ws_planning = wb.active
        ws_planning.title = "Planning Hebdomadaire"
        
        # En-têtes
        ws_planning['A1'] = "PLANNING FRONT OFFICE"
        ws_planning['A1'].font = Font(bold=True, size=14)
        ws_planning['A2'] = f"Semaine du {semaine_debut.strftime('%d/%m/%Y')}"
        ws_planning['A2'].font = Font(bold=True)
        ws_planning['A3'] = "Équipe de 15 personnes : 5 superviseurs + 9 réceptionnistes + 1 concierge"
        ws_planning['A3'].font = Font(italic=True)
        
        # Calcul des dates de la semaine
        dates_semaine = []
        for i in range(7):
            date_jour = semaine_debut + timedelta(days=i)
            dates_semaine.append(date_jour.strftime('%d/%m'))
        
        # Couleurs pour les shifts
        couleur_matin = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        couleur_apres_midi = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")
        couleur_nuit = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        couleur_vide = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        couleur_header = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        couleur_header_jour = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Structure du tableau - En-têtes principales (ligne 5)
        row = 5
        ws_planning.cell(row=row, column=1, value="Employé").font = Font(bold=True)
        ws_planning.cell(row=row, column=1).fill = couleur_header
        ws_planning.cell(row=row, column=2, value="Rôle").font = Font(bold=True)
        ws_planning.cell(row=row, column=2).fill = couleur_header
        ws_planning.cell(row=row, column=3, value="Contrat").font = Font(bold=True)
        ws_planning.cell(row=row, column=3).fill = couleur_header
        
        # En-têtes des jours avec dates (ligne 5)
        col_start = 4
        for i, jour in enumerate(self.jours_semaine):
            date_str = dates_semaine[i]
            col = col_start + (i * 3)
            
            # Fusionner 3 colonnes pour le nom du jour + date
            ws_planning.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
            cell = ws_planning.cell(row=row, column=col, value=f"{jour} {date_str}")
            cell.font = Font(bold=True)
            cell.fill = couleur_header_jour
            cell.alignment = Alignment(horizontal='center')
        
        # Sous-en-têtes des shifts (ligne 6)
        row += 1
        # Colonnes fixes
        ws_planning.cell(row=row, column=1).fill = couleur_header
        ws_planning.cell(row=row, column=2).fill = couleur_header
        ws_planning.cell(row=row, column=3).fill = couleur_header
        
        for i in range(7):  # 7 jours
            col_base = col_start + (i * 3)
            shifts = ["Matin", "AM", "Nuit"]
            for j, shift in enumerate(shifts):
                cell = ws_planning.cell(row=row, column=col_base + j, value=shift)
                cell.font = Font(bold=True, size=10)
                if shift == "Matin":
                    cell.fill = couleur_matin
                elif shift == "AM":
                    cell.fill = couleur_apres_midi
                else:
                    cell.fill = couleur_nuit
                cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Remplissage du planning par employé
        for emp in self.employees:
            # Informations employé
            ws_planning.cell(row=row, column=1, value=f"{emp.prenom} {emp.nom}")
            ws_planning.cell(row=row, column=2, value=emp.role.title())
            ws_planning.cell(row=row, column=3, value=emp.type_contrat.replace('_', ' ').title())
            
            # Coloration de la ligne selon le rôle
            role_color = couleurs_roles.get(emp.role)
            if role_color:
                for col in range(1, 4):  # Colonnes employé, rôle, contrat
                    ws_planning.cell(row=row, column=col).fill = role_color
            
            # Pour chaque jour, remplir les 3 shifts
            for i, jour in enumerate(self.jours_semaine):
                col_base = col_start + (i * 3)
                
                # Matin
                if any(e['prenom'] == emp.prenom for e in planning[jour]['matin']):
                    cell = ws_planning.cell(row=row, column=col_base, value="🌅")
                    cell.fill = couleur_matin
                    cell.font = Font(bold=True)
                else:
                    cell = ws_planning.cell(row=row, column=col_base, value="")
                    cell.fill = couleur_vide
                cell.alignment = Alignment(horizontal='center')
                
                # Après-midi
                if any(e['prenom'] == emp.prenom for e in planning[jour]['apres_midi']):
                    cell = ws_planning.cell(row=row, column=col_base + 1, value="🌆")
                    cell.fill = couleur_apres_midi
                    cell.font = Font(bold=True)
                else:
                    cell = ws_planning.cell(row=row, column=col_base + 1, value="")
                    cell.fill = couleur_vide
                cell.alignment = Alignment(horizontal='center')
                
                # Nuit
                if any(e['prenom'] == emp.prenom for e in planning[jour]['nuit']):
                    cell = ws_planning.cell(row=row, column=col_base + 2, value="🌙")
                    cell.fill = couleur_nuit
                    cell.font = Font(bold=True)
                else:
                    cell = ws_planning.cell(row=row, column=col_base + 2, value="")
                    cell.fill = couleur_vide
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            row += 1
        
        # Ajustement des largeurs de colonnes
        ws_planning.column_dimensions['A'].width = 18  # Employé
        ws_planning.column_dimensions['B'].width = 12  # Rôle
        ws_planning.column_dimensions['C'].width = 15  # Contrat
        
        # Colonnes des shifts (plus petites car juste des icônes)
        for i in range(21):  # 7 jours × 3 shifts = 21 colonnes
            col_letter = chr(68 + i)  # D, E, F, G, ...
            ws_planning.column_dimensions[col_letter].width = 5
        
        # Feuille 2: Validation du Planning
        ws_validation = wb.create_sheet("Validation", 1)
        
        # Titre
        ws_validation['A1'] = "VALIDATION DU PLANNING"
        ws_validation['A1'].font = Font(bold=True, size=12)
        
        # Création des données de validation avec dates
        validation_data = []
        dates_semaine = []
        for i in range(7):
            date_jour = semaine_debut + timedelta(days=i)
            dates_semaine.append(date_jour.strftime('%d/%m'))
            
        for i, jour in enumerate(self.jours_semaine):
            date_str = dates_semaine[i]
            for shift in ['matin', 'apres_midi', 'nuit']:
                equipe = planning[jour][shift]
                nb_superviseurs = len([e for e in equipe if e['role'] == 'superviseur'])
                nb_receptionnistes = len([e for e in equipe if e['role'] == 'receptionniste'])
                nb_concierges = len([e for e in equipe if e['role'] == 'concierge'])
                total = len(equipe)
                
                # Validation des règles
                validation_ok = True
                problemes = []
                
                if shift == 'nuit':
                    if nb_receptionnistes != 2:
                        validation_ok = False
                        problemes.append(f"Doit avoir 2 réceptionnistes (a {nb_receptionnistes})")
                    if nb_superviseurs > 0:
                        validation_ok = False
                        problemes.append("Superviseurs interdits la nuit")
                    if nb_concierges > 0:
                        validation_ok = False
                        problemes.append("Concierge interdit la nuit")
                else:
                    if nb_superviseurs < 1:
                        validation_ok = False
                        problemes.append(f"Doit avoir au moins 1 superviseur (a {nb_superviseurs})")
                    if total > 4:
                        validation_ok = False
                        problemes.append(f"Maximum 4 personnes (a {total})")
                    if jour not in ['Samedi', 'Dimanche'] and shift == 'matin' and nb_concierges != 1:
                        validation_ok = False
                        problemes.append(f"Concierge obligatoire en semaine le matin (a {nb_concierges})")
                    if jour in ['Samedi', 'Dimanche'] and nb_concierges > 0:
                        validation_ok = False
                        problemes.append("Concierge interdit le weekend")
                    if shift == 'apres_midi' and nb_concierges > 0:
                        validation_ok = False
                        problemes.append("Concierge interdit l'après-midi")
                
                validation_data.append({
                    'Jour': f"{jour} {date_str}",
                    'Shift': shift.replace('_', ' ').title(),
                    'Total': total,
                    'Superviseurs': nb_superviseurs,
                    'Réceptionnistes': nb_receptionnistes,
                    'Concierge': nb_concierges,
                    'Statut': 'OK' if validation_ok else 'PROBLÈME',
                    'Détails': ', '.join(problemes) if problemes else 'Conforme'
                })
        
        # En-têtes du tableau de validation
        row = 3
        validation_headers = ['Jour', 'Shift', 'Total', 'Superviseurs', 'Réceptionnistes', 'Concierge', 'Statut', 'Détails']
        for col, header in enumerate(validation_headers, 1):
            cell = ws_validation.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Remplissage des données de validation
        row += 1
        for data in validation_data:
            for col, header in enumerate(validation_headers, 1):
                cell = ws_validation.cell(row=row, column=col, value=data[header])
                # Coloration selon le statut
                if data['Statut'] == 'PROBLÈME':
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif data['Statut'] == 'OK':
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            row += 1
        
        # Ajustement des largeurs pour la validation
        column_widths_validation = [12, 12, 8, 12, 15, 10, 10, 40]
        for i, width in enumerate(column_widths_validation, 1):
            ws_validation.column_dimensions[chr(64 + i)].width = width
        
        # Statistiques de validation
        problemes_count = len([v for v in validation_data if v['Statut'] == 'PROBLÈME'])
        total_shifts = len(validation_data)
        
        row += 2
        ws_validation.cell(row=row, column=1, value="RÉSUMÉ DE VALIDATION").font = Font(bold=True, size=12)
        row += 1
        if problemes_count == 0:
            cell = ws_validation.cell(row=row, column=1, value=f"✅ Planning parfaitement valide ! Tous les {total_shifts} shifts respectent les contraintes.")
            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        else:
            cell = ws_validation.cell(row=row, column=1, value=f"❌ {problemes_count} problème(s) détecté(s) sur {total_shifts} shifts.")
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Feuille 3: Analyse détaillée
        ws_analyse = wb.create_sheet("Analyse")
        
        # Heures par employé
        ws_analyse['A1'] = "ANALYSE DES HEURES PAR EMPLOYÉ"
        ws_analyse['A1'].font = Font(bold=True, size=12)
        
        row = 3
        headers = ['Employé', 'Rôle', 'Contrat', 'Jours Travaillés', 'Jours Contractuels', 'Heures', 'Conforme']
        for col, header in enumerate(headers, 1):
            cell = ws_analyse.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row += 1
        for emp_nom, data in analyse['heures_par_employe'].items():
            ws_analyse.cell(row=row, column=1, value=emp_nom)
            ws_analyse.cell(row=row, column=2, value=data['role'])
            ws_analyse.cell(row=row, column=3, value=data['type_contrat'])
            ws_analyse.cell(row=row, column=4, value=data['jours_travailles'])
            ws_analyse.cell(row=row, column=5, value=data['jours_contractuels'])
            ws_analyse.cell(row=row, column=6, value=data['heures'])
            ws_analyse.cell(row=row, column=7, value="OUI" if data['respect_contrat'] else "NON")
            
            # Coloration selon le rôle
            role_color = couleurs_roles.get(data['role'])
            if role_color:
                for col in range(1, 8):
                    ws_analyse.cell(row=row, column=col).fill = role_color
            
            row += 1
        
        # Violations
        row += 2
        ws_analyse.cell(row=row, column=1, value="VIOLATIONS DE CONTRAINTES").font = Font(bold=True, size=12)
        row += 1
        
        if analyse['violations_contraintes']:
            for violation in analyse['violations_contraintes']:
                ws_analyse.cell(row=row, column=1, value=violation)
                ws_analyse.cell(row=row, column=1).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                row += 1
        else:
            ws_analyse.cell(row=row, column=1, value="Aucune violation détectée ✓")
            ws_analyse.cell(row=row, column=1).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # Statistiques globales
        row += 2
        ws_analyse.cell(row=row, column=1, value="STATISTIQUES GLOBALES").font = Font(bold=True, size=12)
        row += 1
        
        stats = analyse['statistiques_globales']
        ws_analyse.cell(row=row, column=1, value=f"Total shifts semaine: {stats['total_shifts_semaine']}")
        row += 1
        ws_analyse.cell(row=row, column=1, value=f"Total heures semaine: {stats['total_heures_semaine']}")
        row += 1
        ws_analyse.cell(row=row, column=1, value=f"Employés actifs: {stats['nombre_employes_actifs']}")
        
        # Ajustement des largeurs de colonnes pour l'analyse
        for col in range(1, 8):
            ws_analyse.column_dimensions[chr(64 + col)].width = 20
        
        # Feuille 4: Planning individuel (SUPPRIMÉE - déjà dans la feuille principale)
        # Cette feuille n'est plus nécessaire car le format tableau principal 
        # montre déjà le planning individuel de chaque employé
        
        # Sauvegarde en bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()


# ================================
# INITIALISATION PAR DEFAUT
# ================================
def initialiser_equipe_conforme(system):
    """Initialise une équipe conforme aux contraintes du projet - 15 personnes"""
    if system.employees:
        return

    # === 5 SUPERVISEURS (qui font aussi réceptionnistes) ===
    for i in range(1, 6):
        system.ajouter_employe(Employee(
            prenom=f"Superviseur{i}",
            nom=f"Manager{i}",
            type_contrat="temps_plein",
            jours_semaine=5,
            role="superviseur",
            contraintes_speciales={},
            competences=["Management", "Accueil", "Anglais", "Formation"],
            disponible=True,
            motif_indisponibilite="",
            jours_absence=0
        ))

    # === 6 RÉCEPTIONNISTES JOUR ===
    # 4 réceptionnistes temps plein
    for i in range(1, 5):
        system.ajouter_employe(Employee(
            prenom=f"Recep{i}",
            nom=f"Jour{i}",
            type_contrat="temps_plein",
            jours_semaine=5,
            role="receptionniste",
            contraintes_speciales={},
            competences=["Accueil", "Anglais"],
            disponible=True,
            motif_indisponibilite="",
            jours_absence=0
        ))
    
    # 1 réceptionniste mi-temps 4 jours
    system.ajouter_employe(Employee(
        prenom="RecepPT1",
        nom="Partiel4j",
        type_contrat="mi_temps_4j",
        jours_semaine=4,
        role="receptionniste",
        contraintes_speciales={},
        competences=["Accueil", "Anglais"],
        disponible=True,
        motif_indisponibilite="",
        jours_absence=0
    ))
    
    # 1 réceptionniste mi-temps 3 jours
    system.ajouter_employe(Employee(
        prenom="RecepPT2",
        nom="Partiel3j",
        type_contrat="mi_temps_3j",
        jours_semaine=3,
        role="receptionniste",
        contraintes_speciales={},
        competences=["Accueil", "Anglais"],
        disponible=True,
        motif_indisponibilite="",
        jours_absence=0
    ))

    # === 3 RÉCEPTIONNISTES DE NUIT ===
    for i in range(1, 4):
        system.ajouter_employe(Employee(
            prenom=f"Night{i}",
            nom=f"Nuit{i}",
            type_contrat="nuit",
            jours_semaine=5,
            role="receptionniste",
            contraintes_speciales={"horaires": "nuit"},
            competences=["Accueil", "Anglais", "Sécurité"],
            disponible=True,
            motif_indisponibilite="",
            jours_absence=0
        ))

    # === 1 CONCIERGE (off weekend, uniquement matin) ===
    system.ajouter_employe(Employee(
        prenom="Concierge",
        nom="Principal",
        type_contrat="temps_plein",
        jours_semaine=5,
        role="concierge",
        contraintes_speciales={"jours_off": "weekend", "horaires": "matin_uniquement"},
        competences=["Conciergerie", "Anglais", "Tourisme"],
        disponible=True,
        motif_indisponibilite="",
        jours_absence=0
    ))

# ================================
# APPLICATION STREAMLIT
# ================================
def main():
    st.set_page_config(page_title="Planning Front Office Hôtelier", layout="wide")
    st.title("🏨 Système de Planning Front Office Hôtelier")
    st.markdown("**Optimisation des plannings avec contraintes réelles - Équipe de 15 personnes**")

    if 'planning_system' not in st.session_state:
        st.session_state.planning_system = HotelPlanningSystem()

    system = st.session_state.planning_system
    if not system.employees:
        initialiser_equipe_conforme(system)

    # === SIDEBAR ===
    with st.sidebar:
        st.header("⚙️ Configuration")
        st.metric("Capacité hôtel", f"{system.hotel_capacity} chambres")
        st.metric("1 réceptionniste pour", f"{system.clients_per_receptionist} clients")
        st.metric("Max équipe par shift", f"{system.max_receptionists_per_shift} personnes")
        
        st.markdown("---")
        st.markdown("**Composition équipe (15 personnes):**")
        st.metric("Superviseurs", f"{system.nb_supervisors} (font aussi réceptionnistes)")
        st.metric("Réceptionnistes jour", system.nb_receptionnistes_jour)
        st.metric("Réceptionnistes nuit", f"{system.nb_night_receptionists_required}/shift")
        st.metric("Concierge", system.nb_concierges)
        
        st.markdown("---")
        st.markdown("**Contraintes principales:**")
        st.markdown("• 2 jours off/semaine par employé")
        st.markdown("• Max 5 jours consécutifs")
        st.markdown("• Au moins 1 superviseur/shift jour")
        st.markdown("• Concierge: matin seulement, off weekend")
        st.markdown("• 2 réceptionnistes/nuit")

    # === TABS ===
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["👥 Équipe", "📊 Occupation", "📅 Planning", "📈 Analyse", "📥 Export"])

    # === TAB 1: Gestion Équipe ===
    with tab1:
        st.header("👥 Gestion de l'Équipe (15 personnes)")
        
        if system.employees:
            st.subheader("📋 Composition Actuelle")
            
            # Résumé par type avec statuts
            cols = st.columns(4)
            with cols[0]:
                nb_superviseurs = len([e for e in system.employees if e.role == 'superviseur'])
                nb_superviseurs_dispo = len([e for e in system.employees if e.role == 'superviseur' and e.disponible])
                st.metric("Superviseurs", f"{nb_superviseurs_dispo}/{nb_superviseurs}", help="Disponibles/Total")
            with cols[1]:
                nb_recep_jour = len([e for e in system.employees if e.role == 'receptionniste' and e.type_contrat != 'nuit'])
                nb_recep_jour_dispo = len([e for e in system.employees if e.role == 'receptionniste' and e.type_contrat != 'nuit' and e.disponible])
                st.metric("Réceptionnistes jour", f"{nb_recep_jour_dispo}/{nb_recep_jour}", help="Disponibles/Total")
            with cols[2]:
                nb_recep_nuit = len([e for e in system.employees if e.role == 'receptionniste' and e.type_contrat == 'nuit'])
                nb_recep_nuit_dispo = len([e for e in system.employees if e.role == 'receptionniste' and e.type_contrat == 'nuit' and e.disponible])
                st.metric("Réceptionnistes nuit", f"{nb_recep_nuit_dispo}/{nb_recep_nuit}", help="Disponibles/Total")
            with cols[3]:
                nb_concierges = len([e for e in system.employees if e.role == 'concierge'])
                nb_concierges_dispo = len([e for e in system.employees if e.role == 'concierge' and e.disponible])
                st.metric("Concierge", f"{nb_concierges_dispo}/{nb_concierges}", help="Disponibles/Total")
            
            # Vérification de la faisabilité
            employes_disponibles = len([e for e in system.employees if e.disponible])
            total_equipe = len(system.employees)
            
            if employes_disponibles >= 8:  # Minimum viable
                st.success(f"✅ Équipe opérationnelle : {employes_disponibles}/{total_equipe} personnes disponibles")
            elif employes_disponibles >= 5:
                st.warning(f"⚠️ Équipe réduite : {employes_disponibles}/{total_equipe} personnes disponibles - Planning limité possible")
            else:
                st.error(f"❌ Équipe insuffisante : {employes_disponibles}/{total_equipe} personnes disponibles - Planning impossible")
            
            # Tableau détaillé avec statuts et jours d'absence
            df_data = []
            for e in system.employees:
                status_icon = "✅" if e.disponible else "❌"
                if e.disponible and e.jours_absence > 0:
                    status_text = f"Partiellement disponible ({e.jours_travail_max_semaine}/{e.jours_semaine}j) - {e.motif_indisponibilite}"
                elif e.disponible:
                    status_text = f"Disponible ({e.jours_semaine}j)"
                else:
                    status_text = f"Indisponible - {e.motif_indisponibilite}"
                
                df_data.append({
                    'Statut': status_icon,
                    'Prénom': e.prenom,
                    'Nom': e.nom,
                    'Rôle': e.role.title(),
                    'Contrat': e.type_contrat.replace('_', ' ').title(),
                    'Jours Contractuels': e.jours_semaine,
                    'Jours Absence': e.jours_absence if e.disponible else "N/A",
                    'Jours Travail Max': e.jours_travail_max_semaine,
                    'Disponibilité': status_text,
                    'Compétences': ', '.join(e.competences)
                })
            
            df = pd.DataFrame(df_data)
            
            # Style conditionnel pour le tableau
            def highlight_status(row):
                colors = []
                for col in df.columns:
                    if row['Statut'] == '❌':
                        colors.append('background-color: #ffecec; color: #666666')
                    elif row['Jours Travail Max'] < row['Jours Contractuels'] and row['Jours Travail Max'] > 0:
                        colors.append('background-color: #fff8e1; color: #333333')
                    else:
                        colors.append('')
                return colors
            
            styled_df = df.style.apply(highlight_status, axis=1)
            st.dataframe(styled_df, use_container_width=True)

        # Gestion des disponibilités
        with st.expander("📋 Gestion des Disponibilités"):
            st.markdown("**Marquer des employés comme indisponibles (maladie, congés, etc.)**")
            
            # Sélection d'un employé à modifier
            employes_options = [f"{e.prenom} {e.nom} ({e.role}) - {'Disponible' if e.disponible else 'Indisponible'}" for e in system.employees]
            employe_a_modifier_dispo = st.selectbox("Choisir l'employé", options=employes_options, key="modify_availability")
            
            if employe_a_modifier_dispo:
                # Trouver l'employé sélectionné
                prenom_nom_info = employe_a_modifier_dispo.split(' (')[0]
                parts = prenom_nom_info.split(' ')
                if len(parts) >= 2:
                    prenom_sel = parts[0]
                    nom_sel = ' '.join(parts[1:])
                else:
                    prenom_sel = parts[0]
                    nom_sel = ""
                
                # Trouver l'objet employé
                employe_sel = None
                for emp in system.employees:
                    if emp.prenom == prenom_sel and emp.nom == nom_sel:
                        employe_sel = emp
                        break
                
                if employe_sel:
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        nouveau_statut = st.radio(
                            f"Statut de {employe_sel.prenom} {employe_sel.nom}",
                            options=["Disponible", "Partiellement disponible", "Indisponible"],
                            index=0 if employe_sel.disponible and employe_sel.jours_absence == 0 
                                  else (1 if employe_sel.disponible and employe_sel.jours_absence > 0 
                                       else 2),
                            key="availability_status"
                        )
                        
                        # Nombre de jours d'absence si partiellement disponible
                        jours_absence = 0
                        if nouveau_statut == "Partiellement disponible":
                            jours_absence = st.slider(
                                f"Nombre de jours d'absence sur la semaine",
                                min_value=1,
                                max_value=min(6, employe_sel.jours_semaine),
                                value=min(employe_sel.jours_absence if employe_sel.jours_absence > 0 else 1, employe_sel.jours_semaine),
                                key="days_absence"
                            )
                            
                            jours_travail_restants = employe_sel.jours_semaine - jours_absence
                            st.info(f"📊 Jours de travail restants : {jours_travail_restants}/{employe_sel.jours_semaine}")
                    
                    with col2:
                        motif = ""
                        if nouveau_statut in ["Partiellement disponible", "Indisponible"]:
                            motif = st.selectbox(
                                "Motif",
                                options=["Maladie", "Congés payés", "RTT", "Formation", "Congé maternité/paternité", 
                                        "Accident de travail", "Congé sans solde", "Rendez-vous médical", "Autre"],
                                key="unavailability_reason"
                            )
                            
                            if motif == "Autre":
                                motif = st.text_input("Préciser le motif", key="custom_reason")
                        
                        # Affichage des informations contractuelles
                        st.write("**Informations contractuelles :**")
                        st.write(f"• Contrat : {employe_sel.type_contrat.replace('_', ' ').title()}")
                        st.write(f"• Jours contractuels : {employe_sel.jours_semaine} jours/semaine")
                        if nouveau_statut == "Partiellement disponible":
                            st.write(f"• **Disponible : {employe_sel.jours_semaine - jours_absence} jours cette semaine**")
                    
                    col_save, col_reset = st.columns([1, 1])
                    
                    with col_save:
                        if st.button("💾 Mettre à jour le statut", key="update_availability"):
                            if nouveau_statut == "Disponible":
                                employe_sel.disponible = True
                                employe_sel.jours_absence = 0
                                employe_sel.motif_indisponibilite = ""
                                st.success(f"✅ {employe_sel.prenom} {employe_sel.nom} - Disponible ({employe_sel.jours_semaine}j)")
                            elif nouveau_statut == "Partiellement disponible":
                                employe_sel.disponible = True
                                employe_sel.jours_absence = jours_absence
                                employe_sel.motif_indisponibilite = motif
                                jours_restants = employe_sel.jours_semaine - jours_absence
                                st.warning(f"⚠️ {employe_sel.prenom} {employe_sel.nom} - Partiellement disponible ({jours_restants}j) - {motif}")
                            else:  # Indisponible
                                employe_sel.disponible = False
                                employe_sel.jours_absence = 7  # Complètement absent
                                employe_sel.motif_indisponibilite = motif
                                st.error(f"❌ {employe_sel.prenom} {employe_sel.nom} - Indisponible - {motif}")
                            st.rerun()
                    
                    with col_reset:
                        if st.button("🔄 Remettre à 100%", key="reset_to_full"):
                            employe_sel.disponible = True
                            employe_sel.jours_absence = 0
                            employe_sel.motif_indisponibilite = ""
                            st.success(f"✅ {employe_sel.prenom} {employe_sel.nom} remis à 100% disponible")
                            st.rerun()
            
            # Actions de groupe
            st.markdown("---")
            st.markdown("**Actions rapides sur l'équipe :**")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("✅ Tous disponibles", key="all_available"):
                    for emp in system.employees:
                        emp.disponible = True
                        emp.motif_indisponibilite = ""
                    st.success("✅ Toute l'équipe marquée comme disponible")
                    st.rerun()
            
            with col2:
                if st.button("🏖️ Weekend équipe réduite", key="reduced_weekend"):
                    # Garde seulement les superviseurs et réceptionnistes de nuit disponibles
                    for emp in system.employees:
                        if emp.role in ['superviseur'] or (emp.role == 'receptionniste' and emp.type_contrat == 'nuit'):
                            emp.disponible = True
                            emp.motif_indisponibilite = ""
                        else:
                            emp.disponible = False
                            emp.motif_indisponibilite = "Congés"
                    st.info("ℹ️ Mode équipe réduite activé (superviseurs + nuit seulement)")
                    st.rerun()
            
            with col3:
                if st.button("🎯 Occupation faible", key="low_occupation"):
                    # Garde 2-3 superviseurs, 1-2 réceptionnistes jour, 2 nuit, pas de concierge
                    count_superviseurs = 0
                    count_recep_jour = 0
                    for emp in system.employees:
                        if emp.role == 'superviseur' and count_superviseurs < 2:
                            emp.disponible = True
                            emp.motif_indisponibilite = ""
                            count_superviseurs += 1
                        elif emp.role == 'receptionniste' and emp.type_contrat != 'nuit' and count_recep_jour < 1:
                            emp.disponible = True  
                            emp.motif_indisponibilite = ""
                            count_recep_jour += 1
                        elif emp.role == 'receptionniste' and emp.type_contrat == 'nuit':
                            emp.disponible = True
                            emp.motif_indisponibilite = ""
                        else:
                            emp.disponible = False
                            emp.motif_indisponibilite = "Congés - Occupation faible"
                    st.info("ℹ️ Mode occupation faible activé (équipe minimale)")
                    st.rerun()

        with st.expander("➕ Ajouter un Employé"):
            col1, col2 = st.columns(2)
            with col1:
                prenom = st.text_input("Prénom")
                nom = st.text_input("Nom")
                role = st.selectbox("Rôle", ["superviseur", "receptionniste", "concierge"])
            with col2:
                if role == "receptionniste":
                    type_contrat = st.selectbox("Type de contrat", ["temps_plein", "mi_temps_4j", "mi_temps_3j", "nuit"])
                else:
                    type_contrat = st.selectbox("Type de contrat", ["temps_plein"])
                
                competences_base = {
                    "superviseur": ["Management", "Accueil", "Anglais", "Formation"],
                    "receptionniste": ["Accueil", "Anglais"],
                    "concierge": ["Conciergerie", "Anglais", "Tourisme"]
                }
                
                # Liste des compétences de base + langues communes
                competences_disponibles = [
                    "Accueil", "Anglais", "Management", "Conciergerie", "Tourisme", "Sécurité", "Formation",
                    "Français", "Espagnol", "Italien", "Allemand", "Portugais", "Russe", "Chinois", 
                    "Japonais", "Arabe", "Hindi", "Néerlandais", "Suédois", "Norvégien", "Danois"
                ]
                
                # Gestion des langues personnalisées
                if 'langues_personnalisees' not in st.session_state:
                    st.session_state.langues_personnalisees = []
                
                # Toutes les compétences disponibles (base + personnalisées)
                toutes_competences = competences_disponibles + st.session_state.langues_personnalisees
                
                # Interface pour ajouter une langue personnalisée
                col_lang1, col_lang2 = st.columns([3, 1])
                with col_lang1:
                    nouvelle_langue = st.text_input("Ajouter une langue", placeholder="Ex: Coréen, Thaï, Swahili...", key="new_language")
                with col_lang2:
                    if st.button("➕ Ajouter", key="add_language"):
                        if nouvelle_langue and nouvelle_langue not in toutes_competences:
                            st.session_state.langues_personnalisees.append(nouvelle_langue)
                            toutes_competences.append(nouvelle_langue)
                            st.success(f"✅ Langue '{nouvelle_langue}' ajoutée")
                            st.rerun()
                        elif nouvelle_langue in toutes_competences:
                            st.warning("Cette langue existe déjà")
                        else:
                            st.warning("Veuillez saisir une langue")
                
                # Affichage des langues personnalisées ajoutées
                if st.session_state.langues_personnalisees:
                    st.write("**Langues ajoutées :** " + ", ".join(st.session_state.langues_personnalisees))
                    if st.button("🗑️ Effacer toutes les langues ajoutées", key="clear_languages"):
                        st.session_state.langues_personnalisees = []
                        st.success("✅ Langues personnalisées effacées")
                        st.rerun()
                
                competences = st.multiselect("Compétences", 
                    toutes_competences,
                    default=competences_base.get(role, ["Accueil", "Anglais"]))

            if st.button("Ajouter cet Employé"):
                if prenom and nom:
                    system.ajouter_employe(Employee(
                        prenom, nom, type_contrat, 0, role, {}, competences, True, "", 0
                    ))
                    st.success(f"✅ Employé {prenom} {nom} ajouté avec succès!")
                    st.rerun()
                else:
                    st.error("Veuillez remplir le prénom et le nom")

        if system.employees:
            with st.expander("✏️ Modifier un Employé"):
                # Sélection de l'employé à modifier
                employes_options = [f"{e.prenom} {e.nom} ({e.role})" for e in system.employees]
                employe_a_modifier = st.selectbox("Choisir l'employé à modifier", options=employes_options, key="modify_select")
                
                if employe_a_modifier:
                    # Trouver l'employé sélectionné
                    prenom_nom_role = employe_a_modifier.split(' (')[0]
                    parts = prenom_nom_role.split(' ')
                    if len(parts) >= 2:
                        prenom_actuel = parts[0]
                        nom_actuel = ' '.join(parts[1:])
                    else:
                        prenom_actuel = parts[0]
                        nom_actuel = ""
                    
                    # Trouver l'objet employé
                    employe_obj = None
                    for emp in system.employees:
                        if emp.prenom == prenom_actuel and emp.nom == nom_actuel:
                            employe_obj = emp
                            break
                    
                    if employe_obj:
                        st.write(f"**Modification de** : {employe_obj.prenom} {employe_obj.nom} ({employe_obj.role})")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            nouveau_prenom = st.text_input("Nouveau prénom", value=employe_obj.prenom, key="new_prenom")
                            nouveau_nom = st.text_input("Nouveau nom", value=employe_obj.nom, key="new_nom")
                        
                        with col2:
                            # Affichage du rôle et contrat actuels (non modifiables pour conserver la cohérence)
                            st.text_input("Rôle", value=employe_obj.role.title(), disabled=True, key="role_display")
                            st.text_input("Contrat", value=employe_obj.type_contrat.replace('_', ' ').title(), disabled=True, key="contrat_display")
                            
                            # Compétences modifiables avec langues personnalisées
                            if 'langues_personnalisees' not in st.session_state:
                                st.session_state.langues_personnalisees = []
                            
                            competences_disponibles = [
                                "Accueil", "Anglais", "Management", "Conciergerie", "Tourisme", "Sécurité", "Formation",
                                "Français", "Espagnol", "Italien", "Allemand", "Portugais", "Russe", "Chinois", 
                                "Japonais", "Arabe", "Hindi", "Néerlandais", "Suédois", "Norvégien", "Danois"
                            ]
                            toutes_competences = competences_disponibles + st.session_state.langues_personnalisees
                            
                            nouvelles_competences = st.multiselect(
                                "Compétences", 
                                toutes_competences,
                                default=employe_obj.competences,
                                key="new_competences"
                            )
                        
                        col_btn1, col_btn2 = st.columns(2)
                        with col_btn1:
                            if st.button("💾 Sauvegarder les modifications", type="primary", key="save_changes"):
                                if nouveau_prenom and nouveau_nom:
                                    # Mettre à jour l'employé
                                    employe_obj.prenom = nouveau_prenom
                                    employe_obj.nom = nouveau_nom
                                    employe_obj.competences = nouvelles_competences
                                    st.success(f"✅ Employé modifié avec succès : {nouveau_prenom} {nouveau_nom}")
                                    st.rerun()
                                else:
                                    st.error("Veuillez remplir le prénom et le nom")
                        
                        with col_btn2:
                            if st.button("🔄 Annuler", key="cancel_changes"):
                                st.info("Modifications annulées")
                                st.rerun()

        if system.employees:
            with st.expander("🗑️ Supprimer un Employé"):
                noms = [f"{e.prenom} {e.nom} ({e.role})" for e in system.employees]
                to_delete = st.selectbox("Choisir l'employé à supprimer", options=noms)
                if st.button("Confirmer la suppression", type="secondary"):
                    prenom_nom = to_delete.split(' (')[0]
                    prenom, nom = prenom_nom.split(' ', 1)
                    system.supprimer_employe(prenom, nom)
                    st.success(f"✅ {to_delete} supprimé avec succès!")
                    st.rerun()

        with st.expander("🌍 Gestion des Langues"):
            st.markdown("**Langues disponibles dans l'équipe :**")
            
            # Initialisation si nécessaire
            if 'langues_personnalisees' not in st.session_state:
                st.session_state.langues_personnalisees = []
            
            # Collecte de toutes les langues utilisées par l'équipe
            langues_utilisees = set()
            for emp in system.employees:
                for competence in emp.competences:
                    # Identifier les langues (compétences qui ne sont pas techniques)
                    langues_courantes = ["Anglais", "Français", "Espagnol", "Italien", "Allemand", "Portugais", 
                                       "Russe", "Chinois", "Japonais", "Arabe", "Hindi", "Néerlandais", 
                                       "Suédois", "Norvégien", "Danois"] + st.session_state.langues_personnalisees
                    if competence in langues_courantes:
                        langues_utilisees.add(competence)
            
            if langues_utilisees:
                col1, col2 = st.columns(2)
                with col1:
                    st.write("**Langues parlées dans l'équipe :**")
                    for langue in sorted(langues_utilisees):
                        # Compter combien d'employés parlent cette langue
                        nb_locuteurs = sum(1 for emp in system.employees if langue in emp.competences)
                        st.write(f"🗣️ {langue}: {nb_locuteurs} personne(s)")
                
                with col2:
                    st.write("**Statistiques linguistiques :**")
                    st.metric("Langues différentes", len(langues_utilisees))
                    
                    # Langue la plus parlée
                    if langues_utilisees:
                        langue_counts = {}
                        for langue in langues_utilisees:
                            langue_counts[langue] = sum(1 for emp in system.employees if langue in emp.competences)
                        langue_principale = max(langue_counts, key=langue_counts.get)
                        st.metric("Langue principale", f"{langue_principale} ({langue_counts[langue_principale]} pers.)")
            else:
                st.info("Aucune langue spécifique détectée dans l'équipe")
            
            # Gestion des langues personnalisées
            st.markdown("---")
            st.markdown("**Ajouter de nouvelles langues :**")
            
            col_add1, col_add2, col_add3 = st.columns([2, 1, 1])
            with col_add1:
                nouvelle_langue_equipe = st.text_input(
                    "Nouvelle langue", 
                    placeholder="Ex: Coréen, Thaï, Bengali...", 
                    key="new_team_language"
                )
            with col_add2:
                if st.button("➕ Ajouter", key="add_team_language"):
                    if nouvelle_langue_equipe and nouvelle_langue_equipe not in st.session_state.langues_personnalisees:
                        langues_existantes = ["Anglais", "Français", "Espagnol", "Italien", "Allemand", "Portugais", 
                                            "Russe", "Chinois", "Japonais", "Arabe", "Hindi", "Néerlandais", 
                                            "Suédois", "Norvégien", "Danois"]
                        if nouvelle_langue_equipe not in langues_existantes:
                            st.session_state.langues_personnalisees.append(nouvelle_langue_equipe)
                            st.success(f"✅ Langue '{nouvelle_langue_equipe}' ajoutée")
                            st.rerun()
                        else:
                            st.warning("Cette langue existe déjà dans la liste de base")
                    elif nouvelle_langue_equipe in st.session_state.langues_personnalisees:
                        st.warning("Cette langue a déjà été ajoutée")
                    else:
                        st.warning("Veuillez saisir une langue")
            
            with col_add3:
                if st.session_state.langues_personnalisees and st.button("🗑️ Effacer", key="clear_team_languages"):
                    st.session_state.langues_personnalisees = []
                    st.success("✅ Langues personnalisées effacées")
                    st.rerun()
            
            # Affichage des langues personnalisées
            if st.session_state.langues_personnalisees:
                st.write("**Langues ajoutées :** " + ", ".join(st.session_state.langues_personnalisees))

        with st.expander("🔧 Gestion Avancée de l'Équipe"):
            st.markdown("**Actions rapides sur l'équipe :**")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("🔄 Réinitialiser l'équipe par défaut", key="reset_team"):
                    system.employees.clear()
                    initialiser_equipe_conforme(system)
                    st.success("✅ Équipe réinitialisée avec la configuration par défaut")
                    st.rerun()
            
            with col2:
                if st.button("🏥 Absences partielles", key="partial_absences"):
                    # Simule des absences partielles réalistes
                    import random
                    absences_scenarios = [
                        ("Maladie", 1, 2),  # 1-2 jours
                        ("RTT", 1, 1),      # 1 jour
                        ("Rendez-vous médical", 1, 1),  # 1 jour
                        ("Formation", 2, 3),  # 2-3 jours
                        ("Congés payés", 2, 4)  # 2-4 jours
                    ]
                    
                    affected_count = 0
                    for emp in system.employees:
                        if emp.role != 'superviseur' and random.random() < 0.4:  # 40% de chance
                            motif, min_days, max_days = random.choice(absences_scenarios)
                            jours_absence = min(random.randint(min_days, max_days), emp.jours_semaine - 1)
                            if jours_absence > 0:
                                emp.disponible = True
                                emp.jours_absence = jours_absence
                                emp.motif_indisponibilite = motif
                                affected_count += 1
                    
                    st.info(f"ℹ️ {affected_count} employé(s) avec absences partielles simulées")
                    st.rerun()

            with col3:
                if st.button("📊 Scénario réaliste", key="realistic_scenario"):
                    # Scénario réaliste d'un front office
                    scenarios = {
                        "Superviseur1": (True, 0, ""),
                        "Superviseur2": (True, 1, "RTT"),  # 1 jour RTT
                        "Superviseur3": (True, 0, ""),
                        "Recep1": (True, 2, "Congés payés"),  # 2 jours congés
                        "Recep2": (False, 7, "Maladie"),  # Complètement malade
                        "Recep3": (True, 1, "Rendez-vous médical"),  # 1 jour RDV
                        "Night1": (True, 0, ""),
                        "Night2": (True, 1, "Formation"),  # 1 jour formation
                        "Night3": (True, 0, ""),
                        "Concierge": (True, 3, "Congés payés")  # 3 jours congés
                    }
                    
                    for emp in system.employees:
                        if emp.prenom in scenarios:
                            disponible, jours_abs, motif = scenarios[emp.prenom]
                            emp.disponible = disponible
                            emp.jours_absence = jours_abs if disponible else 7
                            emp.motif_indisponibilite = motif
                    
                    st.info("ℹ️ Scénario réaliste appliqué (mix d'absences)")
                    st.rerun()
                # Compteur d'employés par type
                nb_superviseurs = len([e for e in system.employees if e.role == 'superviseur'])
                nb_recep_jour = len([e for e in system.employees if e.role == 'receptionniste' and e.type_contrat != 'nuit'])
                nb_recep_nuit = len([e for e in system.employees if e.role == 'receptionniste' and e.type_contrat == 'nuit'])
                nb_concierges = len([e for e in system.employees if e.role == 'concierge'])
                
                st.write("**Composition actuelle :**")
                st.write(f"👨‍💼 Superviseurs: {nb_superviseurs}")
                st.write(f"👨‍💻 Réceptionnistes jour: {nb_recep_jour}")
                st.write(f"🌙 Réceptionnistes nuit: {nb_recep_nuit}")
                st.write(f"🛎️ Concierges: {nb_concierges}")

    # === TAB 2: Prévisions ===
    with tab2:
        st.header("📊 Prévisions de Check-ins et Check-outs")
        
        # Initialisation des données par défaut
        if 'checkins' not in st.session_state:
            st.session_state.checkins = {
                'Lundi': 180, 'Mardi': 150, 'Mercredi': 200, 'Jeudi': 220,
                'Vendredi': 250, 'Samedi': 300, 'Dimanche': 280
            }
        if 'checkouts' not in st.session_state:
            st.session_state.checkouts = {
                'Lundi': 280, 'Mardi': 300, 'Mercredi': 250, 'Jeudi': 220,
                'Vendredi': 200, 'Samedi': 150, 'Dimanche': 180
            }

        # Boutons de présets
        st.subheader("🎯 Présets d'occupation")
        cols = st.columns(4)
        with cols[0]:
            if st.button("📈 Haute saison"):
                for jour in system.jours_semaine:
                    st.session_state.checkins[jour] = np.random.randint(250, 350)
                    st.session_state.checkouts[jour] = np.random.randint(250, 350)
                st.rerun()
        with cols[1]:
            if st.button("📊 Saison moyenne"):
                for jour in system.jours_semaine:
                    st.session_state.checkins[jour] = np.random.randint(150, 250)
                    st.session_state.checkouts[jour] = np.random.randint(150, 250)
                st.rerun()
        with cols[2]:
            if st.button("📉 Basse saison"):
                for jour in system.jours_semaine:
                    st.session_state.checkins[jour] = np.random.randint(50, 150)
                    st.session_state.checkouts[jour] = np.random.randint(50, 150)
                st.rerun()
        with cols[3]:
            if st.button("🔄 Réinitialiser"):
                st.session_state.checkins = {j: 200 for j in system.jours_semaine}
                st.session_state.checkouts = {j: 200 for j in system.jours_semaine}
                st.rerun()

        # Saisie des données
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("🌅 Check-outs (matin 7h-15h)")
            for jour in system.jours_semaine:
                st.session_state.checkouts[jour] = st.number_input(
                    f"{jour}", 0, 500, st.session_state.checkouts[jour], 
                    key=f"checkout_{jour}", help=f"Nombre de départs le {jour} matin"
                )
        
        with col2:
            st.subheader("🌆 Check-ins (après-midi 15h-23h)")
            for jour in system.jours_semaine:
                st.session_state.checkins[jour] = st.number_input(
                    f"{jour}", 0, 500, st.session_state.checkins[jour], 
                    key=f"checkin_{jour}", help=f"Nombre d'arrivées le {jour} après-midi"
                )

        # Calcul des besoins
        if st.button("🧮 Calculer les Besoins en Personnel", type="primary"):
            besoins = system.calculer_besoins_personnel(st.session_state.checkins, st.session_state.checkouts)
            st.session_state.besoins = besoins
            
            st.subheader("📋 Besoins Calculés")
            
            # Tableau des besoins
            data_besoins = []
            for jour, shifts in besoins.items():
                for shift, infos in shifts.items():
                    if shift == 'nuit':
                        data_besoins.append({
                            'Jour': jour,
                            'Période': shift.replace('_', ' ').title(),
                            'Personnel Total': infos['receptionists'],
                            'Min Superviseurs': infos['superviseurs'],
                            'Concierge': infos['concierge']
                        })
                    else:
                        data_besoins.append({
                            'Jour': jour,
                            'Période': shift.replace('_', ' ').title(),
                            'Personnel Total': infos['total_personnel'],
                            'Min Superviseurs': infos['min_superviseurs'],
                            'Concierge': infos['concierge']
                        })
            
            df_besoins = pd.DataFrame(data_besoins)
            st.dataframe(df_besoins, use_container_width=True)
            
            # Graphique
            fig = px.bar(df_besoins, x='Jour', y=['Personnel Total', 'Min Superviseurs', 'Concierge'], 
                        color='Période', barmode='group',
                        title="Besoins en Personnel par Jour et Période")
            st.plotly_chart(fig, use_container_width=True)
            
            # Résumé
            total_postes = sum(data['Personnel Total'] for data in data_besoins if data['Personnel Total'])
            st.info(f"📊 **Résumé**: {total_postes} postes à pourvoir sur la semaine (hors concierge)")

    # === TAB 3: Planning ===
    with tab3:
        st.header("📅 Génération du Planning")
        
        # Vérifications préalables avec nouvelles conditions
        col1, col2, col3 = st.columns(3)
        with col1:
            besoins_ok = 'besoins' in st.session_state
            st.write("✅ Besoins calculés" if besoins_ok else "❌ Calculez d'abord les besoins")
        with col2:
            employes_disponibles = len([e for e in system.employees if e.disponible])
            equipe_ok = employes_disponibles >= 5  # Minimum viable
            st.write(f"✅ Équipe disponible ({employes_disponibles} pers.)" if equipe_ok else f"❌ Équipe insuffisante ({employes_disponibles} pers.)")
        with col3:
            faisable = st.session_state.get('faisabilite', {}).get('faisable', True)
            st.write("✅ Planning réalisable" if faisable else "❌ Planning impossible")

        if not besoins_ok:
            st.warning("⚠️ Calculez d'abord les besoins dans l'onglet 📊 Occupation.")
        elif not equipe_ok:
            st.error("⚠️ Équipe insuffisante. Il faut au minimum 5 personnes disponibles pour générer un planning.")
            st.info("💡 **Solutions** : Rendez plus d'employés disponibles dans l'onglet 👥 Équipe > Gestion des Disponibilités")
        elif not faisable:
            st.error("⚠️ Planning impossible avec l'équipe disponible. Consultez les problèmes dans l'onglet 📊 Occupation.")
            st.info("💡 **Solutions** : Réduisez l'occupation prévue ou augmentez l'équipe disponible")
        else:
            # Date de début de semaine
            semaine_debut = st.date_input(
                "📅 Date de début de semaine (lundi)",
                value=datetime.now().date(),
                help="Sélectionnez le lundi de la semaine à planifier"
            )
            
            if st.button("✨ Générer le Planning Optimisé", type="primary"):
                with st.spinner("🔄 Génération du planning en cours..."):
                    try:
                        planning = system.generer_planning_optimise(
                            st.session_state.checkins,
                            st.session_state.checkouts,
                            datetime.combine(semaine_debut, datetime.min.time())
                        )
                        
                        if planning:
                            st.session_state.planning = planning
                            st.session_state.semaine_debut = semaine_debut
                            st.success("✅ Planning généré avec succès!")
                        else:
                            st.error("❌ Impossible de générer un planning valide avec les contraintes actuelles.")
                    except Exception as e:
                        st.error(f"❌ Erreur lors de la génération: {str(e)}")

            # Affichage du planning
            if 'planning' in st.session_state and st.session_state.planning:
                st.subheader("📋 Planning Hebdomadaire Généré")
                
                # Calcul des dates de la semaine
                dates_semaine = []
                for i, jour in enumerate(system.jours_semaine):
                    date_jour = semaine_debut + timedelta(days=i)
                    dates_semaine.append(date_jour.strftime('%d/%m'))
                
                # Création du tableau planning avec sous-colonnes
                planning_data = []
                for emp in system.employees:
                    row = {
                        'Employé': f"{emp.prenom} {emp.nom}",
                        'Rôle': emp.role.title(),
                        'Contrat': emp.type_contrat.replace('_', ' ').title()
                    }
                    
                    # Pour chaque jour, créer 3 sous-colonnes (matin/après-midi/nuit)
                    for i, jour in enumerate(system.jours_semaine):
                        date_str = dates_semaine[i]
                        
                        # Matin
                        if any(e['prenom'] == emp.prenom for e in st.session_state.planning[jour]['matin']):
                            row[f"{jour} {date_str} - Matin"] = "🌅 M"
                        else:
                            row[f"{jour} {date_str} - Matin"] = ""
                        
                        # Après-midi
                        if any(e['prenom'] == emp.prenom for e in st.session_state.planning[jour]['apres_midi']):
                            row[f"{jour} {date_str} - AM"] = "🌆 AM"
                        else:
                            row[f"{jour} {date_str} - AM"] = ""
                        
                        # Nuit
                        if any(e['prenom'] == emp.prenom for e in st.session_state.planning[jour]['nuit']):
                            row[f"{jour} {date_str} - Nuit"] = "🌙 N"
                        else:
                            row[f"{jour} {date_str} - Nuit"] = ""
                    
                    planning_data.append(row)
                
                # Affichage du tableau
                df_planning = pd.DataFrame(planning_data)
                
                # Style du tableau avec mise en forme conditionnelle
                def style_planning(val):
                    if "🌅 M" in str(val):
                        return 'background-color: #FFFFCC; color: #000000; font-weight: bold; text-align: center'
                    elif "🌆 AM" in str(val):
                        return 'background-color: #FFCCFF; color: #000000; font-weight: bold; text-align: center'
                    elif "🌙 N" in str(val):
                        return 'background-color: #CCCCFF; color: #000000; font-weight: bold; text-align: center'
                    else:
                        return 'background-color: #F9F9F9; text-align: center'
                
                # Colonnes des shifts (exclure les 3 premières colonnes d'info)
                shift_columns = [col for col in df_planning.columns if any(jour in col for jour in system.jours_semaine)]
                
                # Application du style
                styled_df = df_planning.style.applymap(style_planning, subset=shift_columns)
                
                # Affichage du tableau stylé
                st.dataframe(styled_df, use_container_width=True, height=600)
                
                # Résumé par shift et validation
                st.subheader("📊 Validation du Planning")
                
                validation_data = []
                for i, jour in enumerate(system.jours_semaine):
                    date_str = dates_semaine[i]
                    for shift in ['matin', 'apres_midi', 'nuit']:
                        equipe = st.session_state.planning[jour][shift]
                        nb_superviseurs = len([e for e in equipe if e['role'] == 'superviseur'])
                        nb_receptionnistes = len([e for e in equipe if e['role'] == 'receptionniste'])
                        nb_concierges = len([e for e in equipe if e['role'] == 'concierge'])
                        total = len(equipe)
                        
                        # Validation des règles
                        validation_ok = True
                        problemes = []
                        
                        if shift == 'nuit':
                            if nb_receptionnistes != 2:
                                validation_ok = False
                                problemes.append(f"Doit avoir 2 réceptionnistes (a {nb_receptionnistes})")
                            if nb_superviseurs > 0:
                                validation_ok = False
                                problemes.append("Superviseurs interdits la nuit")
                            if nb_concierges > 0:
                                validation_ok = False
                                problemes.append("Concierge interdit la nuit")
                        else:
                            if nb_superviseurs < 1:
                                validation_ok = False
                                problemes.append(f"Doit avoir au moins 1 superviseur (a {nb_superviseurs})")
                            if total > 4:
                                validation_ok = False
                                problemes.append(f"Maximum 4 personnes (a {total})")
                            if jour not in ['Samedi', 'Dimanche'] and shift == 'matin' and nb_concierges != 1:
                                validation_ok = False
                                problemes.append(f"Concierge obligatoire en semaine le matin (a {nb_concierges})")
                            if jour in ['Samedi', 'Dimanche'] and nb_concierges > 0:
                                validation_ok = False
                                problemes.append("Concierge interdit le weekend")
                            if shift == 'apres_midi' and nb_concierges > 0:
                                validation_ok = False
                                problemes.append("Concierge interdit l'après-midi")
                        
                        validation_data.append({
                            'Jour': f"{jour} {date_str}",
                            'Shift': shift.replace('_', ' ').title(),
                            'Total': total,
                            'Superviseurs': nb_superviseurs,
                            'Réceptionnistes': nb_receptionnistes,
                            'Concierge': nb_concierges,
                            'Statut': '✅ OK' if validation_ok else '❌ Problème',
                            'Détails': ', '.join(problemes) if problemes else 'Conforme'
                        })
                
                df_validation = pd.DataFrame(validation_data)
                
                # Style pour la validation
                def style_validation(row):
                    if '❌' in str(row['Statut']):
                        return ['background-color: #ffcccc'] * len(row)
                    elif '✅' in str(row['Statut']):
                        return ['background-color: #ccffcc'] * len(row)
                    else:
                        return [''] * len(row)
                
                styled_validation = df_validation.style.apply(style_validation, axis=1)
                st.dataframe(styled_validation, use_container_width=True)
                
                # Statistiques de validation
                problemes_count = len([v for v in validation_data if '❌' in v['Statut']])
                total_shifts = len(validation_data)
                
                if problemes_count == 0:
                    st.success(f"✅ Planning parfaitement valide ! Tous les {total_shifts} shifts respectent les contraintes.")
                else:
                    st.error(f"❌ {problemes_count} problème(s) détecté(s) sur {total_shifts} shifts.")
                
                # Vue par équipe pour chaque shift
                with st.expander("👥 Composition détaillée des équipes"):
                    for i, jour in enumerate(system.jours_semaine):
                        date_str = dates_semaine[i]
                        st.write(f"**{jour} {date_str}:**")
                        cols = st.columns(3)
                        
                        shifts_info = [
                            ("🌅 Matin", 'matin'),
                            ("🌆 Après-midi", 'apres_midi'),
                            ("🌙 Nuit", 'nuit')
                        ]
                        
                        for col, (titre, shift_key) in zip(cols, shifts_info):
                            with col:
                                st.write(f"*{titre}:*")
                                equipe = st.session_state.planning[jour][shift_key]
                                if equipe:
                                    for e in equipe:
                                        role_icon = {"superviseur": "👨‍💼", "receptionniste": "👨‍💻", "concierge": "🛎️"}
                                        icon = role_icon.get(e['role'], "👤")
                                        st.write(f"{icon} {e['prenom']} {e['nom']}")
                                else:
                                    st.write("_Aucun employé_")

    # === TAB 4: Analyse ===
    with tab4:
        st.header("📈 Analyse du Planning")
        
        if 'planning' in st.session_state and st.session_state.planning:
            analyse = system.analyser_planning(st.session_state.planning)
            
            # Statistiques globales
            st.subheader("📊 Statistiques Globales")
            stats = analyse['statistiques_globales']
            cols = st.columns(3)
            with cols[0]:
                st.metric("Total shifts", stats['total_shifts_semaine'])
            with cols[1]:
                st.metric("Total heures", stats['total_heures_semaine'])
            with cols[2]:
                st.metric("Employés actifs", f"{stats['nombre_employes_actifs']}/15")
            
            # Analyse par employé
            st.subheader("⏰ Analyse par Employé")
            df_heures = pd.DataFrame.from_dict(analyse['heures_par_employe'], orient='index').reset_index()
            df_heures.rename(columns={'index': 'Employé'}, inplace=True)
            
            # Coloration conditionnelle pour les violations
            def color_violations(row):
                colors = [''] * len(row)
                if not row['respect_contrat']:
                    colors = ['background-color: #ffcccc'] * len(row)
                return colors
            
            st.dataframe(
                df_heures.style.apply(color_violations, axis=1),
                use_container_width=True
            )
            
            # Graphique de répartition des heures
            fig_heures = px.bar(df_heures, x='Employé', y='heures', color='role',
                              title="Répartition des Heures par Employé",
                              color_discrete_map={
                                  'superviseur': '#FFE6CC',
                                  'receptionniste': '#E6F3FF',
                                  'concierge': '#E6FFE6'
                              })
            fig_heures.update_layout(xaxis_tickangle=45)
            st.plotly_chart(fig_heures, use_container_width=True)
            
            # Couverture par shift
            st.subheader("👥 Couverture par Shift")
            df_couverture = pd.DataFrame.from_dict(analyse['couverture_par_shift'], orient='index').reset_index()
            df_couverture.rename(columns={'index': 'Jour_Shift'}, inplace=True)
            
            # Séparer Jour et Shift de manière sécurisée
            split_data = df_couverture['Jour_Shift'].str.split('_', expand=True)
            if split_data.shape[1] >= 2:
                df_couverture['Jour'] = split_data[0]
                df_couverture['Shift'] = split_data[1]
            else:
                # Fallback si le split ne fonctionne pas comme attendu
                df_couverture['Jour'] = df_couverture['Jour_Shift']
                df_couverture['Shift'] = 'unknown'
            
            # Graphique avec détail par rôle
            fig_couverture = px.bar(df_couverture, x='Jour', 
                                  y=['superviseurs', 'receptionnistes', 'concierge'], 
                                  color='Shift',
                                  title="Composition des Équipes par Shift",
                                  barmode='stack')
            st.plotly_chart(fig_couverture, use_container_width=True)
            
            # Détail de la couverture
            with st.expander("📋 Détail de la couverture par shift"):
                for jour in system.jours_semaine:
                    st.write(f"**{jour}:**")
                    for shift in ['matin', 'apres_midi', 'nuit']:
                        equipe_info = analyse['couverture_par_shift'].get(f"{jour}_{shift}", {})
                        st.write(f"  - {shift.title()}: {equipe_info.get('total', 0)} personnes "
                               f"({equipe_info.get('superviseurs', 0)} superviseurs, "
                               f"{equipe_info.get('receptionnistes', 0)} réceptionnistes, "
                               f"{equipe_info.get('concierge', 0)} concierge)")
            
            # Violations de contraintes
            st.subheader("⚠️ Violations de Contraintes")
            violations = analyse['violations_contraintes']
            
            if violations:
                st.error(f"🚨 {len(violations)} violation(s) détectée(s):")
                for i, violation in enumerate(violations, 1):
                    st.write(f"{i}. {violation}")
            else:
                st.success("✅ Aucune violation de contrainte détectée! Le planning respecte toutes les règles.")
        else:
            st.info("📋 Générez d'abord un planning dans l'onglet 📅 Planning pour voir l'analyse.")

    # === TAB 5: Export ===
    with tab5:
        st.header("📥 Export du Planning")
        
        if 'planning' in st.session_state and st.session_state.planning:
            st.success("✅ Planning prêt pour l'export")
            
            # Informations sur l'export
            analyse = system.analyser_planning(st.session_state.planning)
            semaine_debut = st.session_state.get('semaine_debut', datetime.now().date())
            
            col1, col2 = st.columns(2)
            with col1:
                st.info(f"📅 **Semaine du**: {semaine_debut.strftime('%d/%m/%Y')}")
                st.info(f"👥 **Employés actifs**: {analyse['statistiques_globales']['nombre_employes_actifs']}/15")
            with col2:
                st.info(f"⏰ **Total heures**: {analyse['statistiques_globales']['total_heures_semaine']}h")
                st.info(f"🔄 **Total shifts**: {analyse['statistiques_globales']['total_shifts_semaine']}")
            
            # Bouton d'export
            if st.button("📊 Générer le fichier Excel", type="primary"):
                try:
                    with st.spinner("🔄 Génération du fichier Excel..."):
                        excel_data = system.exporter_planning_excel(
                            st.session_state.planning,
                            analyse,
                            datetime.combine(semaine_debut, datetime.min.time())
                        )
                    
                    # Nom du fichier
                    nom_fichier = f"planning_front_office_{semaine_debut.strftime('%Y_%m_%d')}.xlsx"
                    
                    # Bouton de téléchargement
                    st.download_button(
                        label="💾 Télécharger le Planning Excel",
                        data=excel_data,
                        file_name=nom_fichier,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("✅ Fichier Excel généré avec succès!")
                    
                    # Informations sur le contenu du fichier
                    with st.expander("📋 Contenu du fichier Excel"):
                        st.markdown("""
                        **Le fichier Excel contient 3 feuilles:**
                        
                        1. **📅 Planning Hebdomadaire** : Format tableau avec employés en lignes et jours en colonnes
                        2. **✅ Validation** : Vérification automatique de toutes les contraintes par shift
                        3. **📊 Analyse** : Heures par employé, violations de contraintes, statistiques globales
                        
                        **Format du planning principal:**
                        - 📋 **Tableau avec sous-colonnes** : Chaque jour divisé en 3 colonnes (Matin/AM/Nuit)
                        - 📅 **Dates affichées** : Chaque jour avec sa date (ex: "Lundi 15/01")
                        - 🎨 **Icônes visuelles** : 🌅 (matin), 🌆 (après-midi), 🌙 (nuit)
                        - 🎨 **Couleurs par shift** : Jaune (matin), Rose (après-midi), Bleu (nuit)
                        - 👨‍💼 **Distinction des rôles** : Couleurs spécifiques pour employé/rôle/contrat
                        
                        **Caractéristiques de l'équipe:**
                        - 🏨 **15 personnes au total**
                        - 👨‍💼 **5 superviseurs** (font aussi office de réceptionnistes)
                        - 👨‍💻 **9 réceptionnistes** (6 jour + 3 nuit)
                        - 🛎️ **1 concierge** (5j/7, matin uniquement)
                        
                        **Fonctionnalités:**
                        - 🎨 Mise en forme professionnelle avec couleurs
                        - ✅ Validation automatique de toutes les contraintes
                        - 📈 Analyse détaillée des heures et de la conformité
                        - 📋 Format tableau facile à lire et imprimer
                        """)
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors de la génération du fichier Excel: {str(e)}")
            
            # Aperçu des violations
            violations = analyse['violations_contraintes']
            if violations:
                st.warning(f"⚠️ Attention: {len(violations)} violation(s) de contrainte dans le planning")
                with st.expander("Voir les violations"):
                    for violation in violations:
                        st.write(f"• {violation}")
            else:
                st.success("✅ Le planning respecte toutes les contraintes")
                
        else:
            st.info("📋 Générez d'abord un planning dans l'onglet 📅 Planning pour pouvoir l'exporter.")
            
            # Aide pour l'export
            with st.expander("ℹ️ Comment utiliser l'export"):
                st.markdown("""
                **Pour exporter votre planning:**
                
                1. 👥 Configurez votre équipe dans l'onglet "Équipe" (15 personnes)
                2. 📊 Saisissez les prévisions d'occupation dans l'onglet "Occupation"
                3. 📅 Générez le planning dans l'onglet "Planning"
                4. 📥 Revenez ici pour télécharger le fichier Excel
                
                **Le fichier Excel contiendra:**
                - Planning complet avec mise en forme
                - Analyse détaillée des heures par employé
                - Vérification de toutes les contraintes
                - Planning individuel pour chaque membre de l'équipe
                - Composition optimale : superviseurs + réceptionnistes selon les besoins
                """)

if __name__ == "__main__":
    main()
